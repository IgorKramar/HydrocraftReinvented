# -*- coding: utf-8 -*-
"""model.json -> HCR_Database.xlsx (Сводка, Предметы, Рецепты, Книги)."""
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from hcr_paths import DOCS, load_model, mod_version, require_vanilla_names

M = load_model()
require_vanilla_names(M)
items, recipes = M["items"], M["recipes"]
ru_items, ru_recipes = M["ru_items"], M["ru_recipes"]
van_ru, van_en = M["van_ru"], M["van_en"]
book_teaches, producers, consumers = M["book_teaches"], M["producers"], M["consumers"]

def ru_item(fid):
    if fid in ru_items: return ru_items[fid]
    short = fid.split(".", 1)[-1]
    return van_ru.get(fid) or van_ru.get(short) or ""

def en_item(fid):
    if fid in items: return items[fid]["en"]
    short = fid.split(".", 1)[-1]
    return van_en.get(fid) or van_en.get(short) or short

def disp(fid):
    return ru_item(fid) or en_item(fid)

ITEM_LINE = re.compile(r'^item\s+(\d+)\s+(.*)$')
def pretty_lines(lines):
    out = []
    for line in lines:
        m = ITEM_LINE.match(line)
        if not m:
            out.append(line)  # fluid/energy как есть
            continue
        cnt, rest = m.group(1), m.group(2).strip()
        if rest.startswith("tags["):
            tags = re.sub(r'^tags\[([^\]]*)\].*$', r'\1', rest)
            keep = " (инструмент)" if "mode:keep" in rest else ""
            out.append(f"{cnt}× инструмент с тегом [{tags}]{keep}")
            continue
        lm = re.match(r'^\[([^\]]*)\]', rest)
        ids = [x.strip() for x in lm.group(1).split(";")] if lm else [rest.split()[0]]
        names = " / ".join(dict.fromkeys(disp(i) for i in ids if i))
        keep = " (не расходуется)" if "mode:keep" in rest else ""
        out.append(f"{cnt}× {names}{keep}")
    return "; ".join(out)

def out_ids(r):
    ids = []
    for line in r["outputs"]:
        m = ITEM_LINE.match(line)
        if m:
            rest = m.group(2).strip()
            lm = re.match(r'^\[([^\]]*)\]', rest)
            ids += [x.strip() for x in lm.group(1).split(";")] if lm else [rest.split()[0]]
    return [i for i in ids if i]

# книги: id -> учит N рецептов
book_recipes = defaultdict(list)
for rn, books in book_teaches.items():
    for b in books:
        book_recipes[b].append(rn)

HDR = Font(name="Arial", bold=True, color="FFFFFF", size=10)
FILL = PatternFill("solid", fgColor="4F6228")
BASE = Font(name="Arial", size=10)

def style_sheet(ws, widths):
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for cell in ws[1]:
        cell.font = HDR; cell.fill = FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

wb = Workbook()

# ---------- Сводка ----------
ws = wb.active; ws.title = "Сводка"
rows = [
    ["Hydrocraft Reinvented [B42] — сводная база данных", ""],
    [f"Источник: common/media/scripts + Translate/RU (версия мода {mod_version()})", ""],
    ["", ""],
    ["Предметов", "=COUNTA(Предметы!A:A)-1"],
    ["Рецептов", "=COUNTA(Рецепты!A:A)-1"],
    ["Рецептов, требующих изучения", '=COUNTIF(Рецепты!G:G,"да")'],
    ["Рецептов, изучаемых по книгам", '=COUNTIF(Рецепты!H:H,"<>—")-1'],
    ["Книг и чертежей с рецептами", "=COUNTA(Книги!A:A)-1"],
]
for r_i, (a, b) in enumerate(rows, 1):
    ws.cell(r_i, 1, a).font = Font(name="Arial", size=12, bold=(r_i == 1))
    if b:
        c = ws.cell(r_i, 2, b); c.font = Font(name="Arial", size=12, bold=True)
ws.column_dimensions["A"].width = 44; ws.column_dimensions["B"].width = 14

# ---------- Предметы ----------
ws = wb.create_sheet("Предметы")
ws.append(["Идентификатор", "Название (RU)", "Название (EN)", "Класс", "Категория",
           "Крафтится рецептами", "Используется в рецептах", "Файл"])
for fid in sorted(items):
    it = items[fid]
    ws.append([fid, ru_item(fid), it["en"], it["type"].replace("base:", ""), it["cat"],
               len(producers.get(fid, [])), len(consumers.get(fid, [])), it["file"]])
style_sheet(ws, [34, 38, 34, 12, 12, 11, 11, 24])

# ---------- Рецепты ----------
ws = wb.create_sheet("Рецепты")
ws.append(["Рецепт (EN)", "Рецепт (RU)", "Категория", "Время", "Навык", "Опыт",
           "Нужно изучить", "Книга", "Ингредиенты", "Результат", "Выход идёт дальше в"])
for rn in sorted(recipes):
    r = recipes[rn]
    books = "; ".join(disp(b) for b in book_teaches.get(rn, [])) or "—"
    oids = out_ids(r)
    downstream = sorted({c for o in oids for c in consumers.get(o, []) if c != rn})
    down = "; ".join((ru_recipes.get(d) or d) for d in downstream[:6])
    if len(downstream) > 6:
        down += f" … (всего {len(downstream)})"
    ws.append([rn, ru_recipes.get(rn, ""), r["cat"], int(r["time"]) if r["time"].isdigit() else r["time"],
               r["skill"], r["xp"], "да" if r["learn"] else "нет", books,
               pretty_lines(r["inputs"]), pretty_lines(r["outputs"]), down or "—"])
style_sheet(ws, [36, 40, 14, 8, 16, 14, 9, 30, 60, 34, 50])

# ---------- Книги ----------
ws = wb.create_sheet("Книги")
ws.append(["Идентификатор", "Книга (RU)", "Книга (EN)", "Рецептов", "Учит рецептам (RU)"])
for b in sorted(book_recipes):
    rl = sorted(book_recipes[b])
    ws.append([b, ru_item(b), en_item(b), len(rl),
               "; ".join((ru_recipes.get(x) or x) for x in rl)])
style_sheet(ws, [30, 34, 30, 10, 120])

for sheet in wb.worksheets:
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.font is None or not cell.font.bold:
                cell.font = BASE

DOCS.mkdir(exist_ok=True)
wb.save(DOCS / "HCR_Database.xlsx")
print("saved", DOCS / "HCR_Database.xlsx")
